import csv
from typing import List
import io

import openpyxl

from fastapi import HTTPException
from fastapi.datastructures import UploadFile

from trucost.core.models.athena_query import AthenaQueryCreate


async def parse_csv_file(file: UploadFile) -> List[AthenaQueryCreate]:
    queries = []
    try:
        content = await file.read()

        # Use StringIO to handle newlines properly
        csv_file = io.StringIO(content.decode("utf-8"))
        reader = csv.DictReader(csv_file)
        for row in reader:
            queries.append(AthenaQueryCreate(**row))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error parsing CSV file: {str(e)}")
    finally:
        await file.seek(0)
    return queries


def parse_excel_file(
    file: UploadFile, sheet_name: str | None = None
) -> List[AthenaQueryCreate]:
    """Parse an Excel file and return a list of AthenaQueryCreate objects"""

    COL_DB_ATTR_MAP = {
        "Master Category": "category",
        "Master Type": "category_type",
        "Query Type": "query_type",
        "Query Sub Type": "query_subtype",
        "Query (Legasy CUR)": "query",
    }

    queries = []
    try:
        content = file.file.read()

        # Use BytesIO to handle newlines properly
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        sheet = workbook.active if sheet_name is None else workbook[sheet_name]

        idx_col_map = {i: col.value for i, col in enumerate(sheet["1"])}
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            queries.append(
                AthenaQueryCreate(
                    **{
                        COL_DB_ATTR_MAP[idx_col_map[i]]: v
                        for i, v in enumerate(row)
                        if i in idx_col_map
                    }
                )
            )
    except Exception as e:
        print(f"Error parsing Excel file: {str(e)}")
        import traceback

        traceback.print_exc()
        raise HTTPException(
            status_code=400, detail=f"Error parsing Excel file: {str(e)}"
        )
    return queries
